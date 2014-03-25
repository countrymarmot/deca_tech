# ShiftDataImporter.py
# Created by Craig Bishop on 11 March 2013
#
# pyrite
# Copyright 2013 All Rights Reserved
#

from PySide.QtGui import QDialog, QFileDialog, QMessageBox,\
    QInputDialog
from Ui_ExcelShiftDataSelector import Ui_ExcelShiftDataSelector
import re
from malachite import zinc_pb2
from google.protobuf import text_format
import openpyxl
from Project import Project
import math


def importShiftData(parent):
  pyrite_file_name, filt = QFileDialog.getOpenFileName(parent,
      "Select Pyrite Project", filter="Pyrite Projects (*.pyrite)")
  if not pyrite_file_name:
    return
  pyrite_project = Project.fromFile(pyrite_file_name)
  design_package = pyrite_project.design.package
  design_die = design_package.dies[0]

  excel_file_name, filt = QFileDialog.getOpenFileName(parent,
      "Select Shifts Excel File", filter="Excel Files (*.xlsx *.xlsm)")
  if not excel_file_name:
    return

  panel_id = ""
  match = re.search("(?:D[0-9]+_[A-Z]_)([0-9]+)(?:.*\.xls.*)|(?:MP.+\.)"
      "([0-9]+)(?:\..*\.xls.*)", excel_file_name)
  if match:
    panel_id += [g for g in match.groups() if g is not None][0]

  panel_id, good = QInputDialog.getText(parent, "Panel ID",
      "Enter the panel ID", text=panel_id)
  if not panel_id:
    return

  excel_wb = openpyxl.reader.excel.load_workbook(excel_file_name)
  sheet_names = excel_wb.get_sheet_names()

  preselect_sheets = [s for s in sheet_names
                      if "calculated values" in s.lower()]
  global_offset_txt = ""
  if preselect_sheets:
    sheet = excel_wb.get_sheet_by_name(preselect_sheets[0])
    if sheet.cell("A1").value == "Global X Shift:" and\
            sheet.cell("A2").value == "Global Y Shift:":
              global_offset_txt = "{0}, {1}".format(sheet.cell("C1").value,
                  sheet.cell("C2").value)

  while True:
    global_offset_txt, good = QInputDialog.getText(parent, "Global Offset",
        "Enter the panel's global offset (x, y) (e.g. 533.33, 434.3)",
        text=global_offset_txt)
    if not global_offset_txt:
      return
    try:
      tokens = [t.strip() for t in global_offset_txt.split(',')]
      global_x = float(tokens[0])
      global_y = float(tokens[1])
      break
    except:
      QMessageBox.warning(parent, "Global Offset",
          "{0} is not a valid global offset".format(global_offset_txt))

  data_selector = ExcelShiftDataSelector(parent, sheet_names)
  data_selector.setWindowTitle("Select Die Shifts")
  preselect_sheets = [s for s in sheet_names
                      if "calculated values" in s.lower()]
  if preselect_sheets:
    data_selector.cmboSheet.setCurrentIndex(
        sheet_names.index(preselect_sheets[0]))
    sheet = excel_wb.get_sheet_by_name(preselect_sheets[0])
    if sheet.cell("B4").value == "Package Nominal":
      if sheet.cell("B5").value == "X (um)":
        data_selector.txtNomX.setText("B6")
      if sheet.cell("C5").value == "Y (um)":
        data_selector.txtNomY.setText("C6")
    if sheet.cell("D4").value == "Die Shift (Deviation)":
      if sheet.cell("D5").value == "X (um)":
        data_selector.txtXShift.setText("D6")
      if sheet.cell("E5").value == "Y (um)":
        data_selector.txtYShift.setText("E6")
      if sheet.cell("F5").value == "Theta (deg)":
        data_selector.txtTheta.setText("F6")
    if sheet.cell("D1").value == "Package Count:":
      data_selector.txtCount.setText(str(sheet.cell("F1").value))

  data_selector.exec_()
  if data_selector.result() == QDialog.Rejected:
    return
  die_shifts = []
  worksheet = excel_wb.get_sheet_by_name(data_selector.sheetName)
  for i in range(data_selector.count):
    nomx = float(worksheet.cell(row=data_selector.nomXRow + i,
        column=data_selector.nomXCol).value)
    nomy = float(worksheet.cell(row=data_selector.nomYRow + i,
        column=data_selector.nomYCol).value)
    shiftx = float(worksheet.cell(row=data_selector.shiftXRow + i,
        column=data_selector.shiftXCol).value)
    shifty = float(worksheet.cell(row=data_selector.shiftYRow + i,
        column=data_selector.shiftYCol).value)
    theta = float(worksheet.cell(row=data_selector.thetaRow + i,
        column=data_selector.thetaCol).value)
    die_shifts.append((nomx, nomy, shiftx, shifty, theta))

  data_selector = ExcelShiftDataSelector(parent, sheet_names)
  data_selector.setWindowTitle("Select Fiducial Die Shifts")
  data_selector.txtDieName.setText("FIDUCIAL")
  preselect_sheets = [s for s in sheet_names
                      if "fiducial" in s.lower()]
  if preselect_sheets:
    data_selector.cmboSheet.setCurrentIndex(
        sheet_names.index(preselect_sheets[0]))
  data_selector.exec_()
  if data_selector.result() == QDialog.Rejected:
    return
  fiducial_shifts = []
  worksheet = excel_wb.get_sheet_by_name(data_selector.sheetName)
  for i in range(data_selector.count):
    nomx = float(worksheet.cell(row=data_selector.nomXRow + i,
        column=data_selector.nomXCol).value)
    nomy = float(worksheet.cell(row=data_selector.nomYRow + i,
        column=data_selector.nomYCol).value)
    shiftx = float(worksheet.cell(row=data_selector.shiftXRow + i,
        column=data_selector.shiftXCol).value)
    shifty = float(worksheet.cell(row=data_selector.shiftYRow + i,
        column=data_selector.shiftYCol).value)
    theta = float(worksheet.cell(row=data_selector.thetaRow + i,
        column=data_selector.thetaCol).value)
    fiducial_shifts.append((nomx, nomy, shiftx, shifty, theta))

  shifts = zinc_pb2.Shifts()
  shifts.designNumber = design_package.designNumber
  shifts.designRevision = design_package.designRevision
  shifts.panelID = panel_id
  shifts.globalOffset.x = global_x
  shifts.globalOffset.y = global_y

  constraints = design_die.shiftConstraints

  for die_unit in die_shifts:
    zunit = shifts.units.add()
    zunit.number = die_shifts.index(die_unit)
    zunit.center.x = die_unit[0]
    zunit.center.y = die_unit[1]
    zdieshift = zunit.die.add()
    zdieshift.name = design_die.name
    zdieshift.nominalXY.x = die_unit[0] + design_die.center.x
    zdieshift.nominalXY.y = die_unit[1] + design_die.center.y
    zdieshift.shift.x = die_unit[2]
    zdieshift.shift.y = die_unit[3]
    zdieshift.theta = math.radians(die_unit[4])
    zunit.inSpec = check_in_spec(zdieshift, constraints)

  for fiducial_unit in fiducial_shifts:
    zunit = shifts.referenceUnits.add()
    zunit.number = fiducial_shifts.index(fiducial_unit)
    zunit.center.x = die_unit[0]
    zunit.center.y = die_unit[1]
    zdieshift = zunit.die.add()
    zdieshift.name = "FIDUCIAL"
    zdieshift.nominalXY.x = fiducial_unit[0]
    zdieshift.nominalXY.y = fiducial_unit[1]
    zdieshift.shift.x = fiducial_unit[2]
    zdieshift.shift.y = fiducial_unit[3]
    zdieshift.theta = math.radians(fiducial_unit[4])
    zunit.inSpec = True

  filename, filt = QFileDialog.getSaveFileName(parent, "Save Shifts File",
      filter="Onyx Shift Files (*.onyxshifts)")
  if not filename:
    return
  file = open(filename, "wb")
  file.write(text_format.MessageToString(shifts))
  file.close()


def check_in_spec(zdieshift, constraints):
  return (zdieshift.shift.x >= constraints.minX and
          zdieshift.shift.x <= constraints.maxX and
          zdieshift.shift.y >= constraints.minY and
          zdieshift.shift.y <= constraints.maxY and
          zdieshift.theta >= constraints.minTheta and
          zdieshift.theta <= constraints.maxTheta)


class ExcelShiftDataSelector(QDialog, Ui_ExcelShiftDataSelector):
  def __init__(self, parent, sheetNames):
    QDialog.__init__(self)
    self.setupUi(self)
    self.connectActions()
    self.setupSheets(sheetNames)
    self.nomXRow = 0
    self.nomXCol = 0
    self.nomYRow = 0
    self.nomYCol = 0
    self.shiftXRow = 0
    self.shiftXCol = 0
    self.shiftYRow = 0
    self.shiftYCol = 0
    self.thetaRow = 0
    self.thetaCol = 0
    self.sheetName = ""
    self.count = 0
    self.dieName = ""

  def connectActions(self):
    self.btnOK.clicked.connect(self.OK)
    self.btnCancel.clicked.connect(self.reject)

  def setupSheets(self, sheetNames):
    for name in sheetNames:
      self.cmboSheet.addItem(name)

  def columnForName(self, colName):
    colName = colName.lower()
    alphabet = 'abcdefghijklmnopqrstuvwxyz'
    multiplier = 0
    col = 0
    for c in reversed(colName):
      col += (alphabet.index(c) + 1) * max(1, multiplier * 26)
      multiplier += 1
    return col

  def rowAndColumnForCellName(self, cellName):
    grps = re.search('([a-zA-Z]*)([0-9]*)', cellName)
    colName = grps.group(1)
    col = self.columnForName(colName) - 1
    row = int(grps.group(2)) - 1
    return row, col

  def OK(self):
    if not str(self.txtNomX.text()):
      return
    if not str(self.txtNomY.text()):
      return
    if not str(self.txtXShift.text()):
      return
    if not str(self.txtYShift.text()):
      return
    if not str(self.txtTheta.text()):
      return
    if not str(self.txtCount.text()):
      return
    if not str(self.txtDieName.text()):
      return

    self.nomXRow, self.nomXCol =\
        self.rowAndColumnForCellName(self.txtNomX.text())
    self.nomYRow, self.nomYCol =\
        self.rowAndColumnForCellName(self.txtNomY.text())
    self.shiftXRow, self.shiftXCol =\
        self.rowAndColumnForCellName(self.txtXShift.text())
    self.shiftYRow, self.shiftYCol =\
        self.rowAndColumnForCellName(self.txtYShift.text())
    self.thetaRow, self.thetaCol =\
        self.rowAndColumnForCellName(self.txtTheta.text())
    self.sheetName = self.cmboSheet.currentText()
    self.count = int(self.txtCount.text())
    self.dieName = self.txtDieName.text()
    self.accept()
